import os
import textwrap

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

from parse import course_parse as course_parse_handler
from parse import authorization as authorization_handler


def create_excel_count(wb, name, course_data):
    ws = wb.active
    next_row = ws.max_row + 2
    module_elements = dict()
    resource_types = ['Файл', 'Задание', 'Тест', 'Ссылка', 'Страница']

    for module in course_data:
        module_elements[module] = {}
        for resource_type in resource_types:
            module_elements[module][resource_type] = 0

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=len(course_data))
    cell = ws.cell(row=next_row, column=1, value=name)
    cell.fill = PatternFill('solid', fgColor='FFb3ffe6')
    cell.border = thin_border

    for col, module in enumerate(course_data, start=1):
        row = next_row + 2
        ws.cell(row=next_row + 1, column=col, value=module).border = thin_border
        ws.column_dimensions[ws.cell(row=next_row + 1, column=col).column_letter].width = len(module) + 4 if len(
            module) < 23 else 26
        for value in course_data[module]:
            if value == 'Folder':
                for child in course_data[module]['Folder']:
                    for child_value in course_data[module]['Folder'][child]:
                        module_elements[module]['Файл'] += 1
            else:
                cdmv = course_data[module][value]
                if cdmv in resource_types:
                    module_elements[module][cdmv] += 1
        for key, value in module_elements[module].items():
            ws.cell(row=row, column=col, value=f'{key}: {value}').border = thin_border
            row += 1
    return wb


def get_excel(specialty, data):
    # authorization_handler()
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=specialty)
    for course in data['data']:
        course_name = course
        course_data = course_parse_handler(data['data'][course]['id'])
        create_excel_count(wb, course_name, course_data)
    return wb
